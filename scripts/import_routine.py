"""Turns the family routine workbook into what NIDO runs.

Two outputs, on purpose:

  examples/abril/<date>.json   one scenario fixture per day - only what the engine may look at:
                               when things happen, how long they take, what depends on what.
  nido-plan.json               the same thing as one file to carry to a phone: the fixtures plus
                               everything the engine must NOT look at - what to do, what to say
                               when she refuses, the food, the safety note, the weaning stage and
                               the sources.

The bundle is never published. It is loaded once into the app on each phone and stays there: a
child routine with feeding, sleep and weaning detail is not something to leave on a public site,
and the app is built so it never has to be.

Keeping them apart is the whole point. The engine decides times and nothing else; the guidance is
the caregiver's own plan, quoted back, never invented here.

    python import_routine.py <workbook.xlsx> <repo-root>
"""
import json, os, re, sys, unicodedata
import openpyxl

WORKBOOK, ROOT = sys.argv[1], sys.argv[2]

# type -> (category, priority, minutes, timing style, slack before/after)
# Slack is the guardrail the engine gets: how far it may move something before it has to say so.
KIND = {
    "DESPERTAR":       ("personal",     "P1", 10, "anchor", 30, 45),
    "MOVIMIENTO":      ("development",  "P4", 10, "window", 15, 45),
    "DESAYUNO":        ("feeding",      "P1", 25, "anchor", 20, 40),
    "FIN COMIDA":      (None, None, 0, None, 0, 0),          # folded into the meal it closes
    "CONEXIÓN":        ("development",  "P3", 15, "window", 20, 45),
    "JUEGO ACTIVO":    ("development",  "P4", 30, "window", 20, 60),
    "LENGUAJE":        ("development",  "P4", 20, "window", 20, 60),
    "SNACK 1":         ("feeding",      "P2", 15, "anchor", 20, 40),
    "PRE-SIESTA":      ("prep",         "P2", 15, "window", 15, 30),
    "SIESTA 1":        ("sleep",        "P2", 45, "window", 15, 40),
    "JUEGO FINO":      ("development",  "P4", 20, "window", 20, 60),
    "PREPARAR COMIDA": ("prep",         "P3", 15, "window", 20, 30),
    "COMIDA":          ("feeding",      "P1", 30, "anchor", 20, 45),
    "QUIET PLAY":      ("development",  "P4", 20, "window", 20, 60),
    "EXTERIOR":        ("outdoor",      "P3", 45, "window", 25, 60),
    "BAJAR RITMO":     ("prep",         "P3", 10, "window", 15, 30),
    "SIESTA 2":        ("sleep",        "P2", 70, "dependent", 0, 0),
    "SNACK 2":         ("feeding",      "P2", 15, "anchor", 20, 40),
    "JUEGO SOCIAL":    ("development",  "P4", 25, "window", 25, 60),
    "CENA":            ("feeding",      "P1", 30, "anchor", 20, 40),
    "CALMA":           ("development",  "P3", 20, "window", 15, 40),
    "BAÑO":            ("hygiene",      "P2", 15, "window", 15, 30),
    "PECHO BEDTIME":   ("breastfeeding","P2", 15, "window", 15, 30),
    "DIENTES":         ("hygiene",      "P2",  5, "window", 10, 25),
    "LIBROS":          ("development",  "P3", 15, "window", 10, 25),
    "CAMA":            ("sleep",        "P1", 30, "anchor", 20, 30),
}

# The workbook says it in prose: "si la siesta 2 fue corta, bedtime 15-30 min antes". The engine
# already knows how to do that, so it is stated here as policy rather than left to a human to
# remember at 7pm.
BEDTIME_POLICY = {"type": "durationResponsive", "reference": "siesta-2", "shortfallMinutes": 25, "shiftMinutes": 20}

# What the caregiver should read on the strip. The workbook names steps by what to do ("Dormir",
# "Panal + agua"), which repeats three times a day; the kind is what tells them apart.
LABEL = {
    "DESPERTAR": "Despertar", "MOVIMIENTO": "Piso activo", "DESAYUNO": "Desayuno",
    "CONEXION": "Conexión", "JUEGO ACTIVO": "Juego activo", "LENGUAJE": "Lenguaje",
    "SNACK 1": "Snack 1", "SNACK 2": "Snack 2", "PRE-SIESTA": "Antes de la siesta",
    "SIESTA 1": "Siesta 1", "SIESTA 2": "Siesta 2", "JUEGO FINO": "Juego fino",
    "PREPARAR COMIDA": "Preparar la comida", "COMIDA": "Comida", "QUIET PLAY": "Juego tranquilo",
    "EXTERIOR": "Salir", "BAJAR RITMO": "Bajar el ritmo", "JUEGO SOCIAL": "Juego social",
    "CENA": "Cena", "CALMA": "Calma", "BANO": "Baño", "PECHO BEDTIME": "Pecho de la noche",
    "DIENTES": "Dientes", "LIBROS": "Libros", "CAMA": "Dormir",
}


def label_for(kind, seen):
    """Two naps means two 'Antes de la siesta'. Number them rather than repeat them."""
    base = LABEL.get(unicodedata.normalize("NFKD", kind).encode("ascii", "ignore").decode(), kind.title())
    count = seen.get(base, 0) + 1
    seen[base] = count
    return base if count == 1 else "%s %d" % (base, count)


def slug(text):
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def clock(value):
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    return str(value).strip()[:5]


def shift(hhmm, minutes):
    hour, minute = (int(p) for p in hhmm.split(":"))
    total = max(0, min(23 * 60 + 59, hour * 60 + minute + minutes))
    return "%02d:%02d" % (total // 60, total % 60)


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text not in {"—", "-"} else None


wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

# ---------------------------------------------------------------- the days
rows_by_date = {}
weekday_by_date = {}
ws = wb["RUTINA DETALLADA"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if not row or not row[0] or not row[2]:
        continue
    date = row[0].strftime("%Y-%m-%d") if hasattr(row[0], "strftime") else str(row[0])[:10]
    rows_by_date.setdefault(date, []).append(row)
    weekday_by_date[date] = row[1]

# ------------------------------------------------------------ weaning stage
stages = {}
ws = wb["DESTETE GRADUAL"]
for row in ws.iter_rows(min_row=4, max_row=10, values_only=True):
    if not row or not row[0]:
        continue
    label = str(row[0])                                   # "Domingo 16-Aug"
    stages[label.split()[0].lower()] = {
        "label": label,
        "stage": clean(row[1]),
        "goal": clean(row[2]),
        "morning": clean(row[3]),
        "nap1": clean(row[4]),
        "nap2": clean(row[5]),
        "bedtime": clean(row[6]),
        "night": clean(row[7]),
        "advance": clean(row[8]),
    }

# ------------------------------------------------------------------- menus
menus = {}
ws = wb["MENÚ + PORCIONES"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if not row or not row[0] or not row[2]:
        continue
    menus.setdefault(str(row[0]).lower(), {})[clock(row[2])] = {
        "moment": clean(row[1]),
        "menu": clean(row[3]),
        "portion": clean(row[4]),
        "density": clean(row[5]),
        "drink": clean(row[6]),
        "ifMore": clean(row[7]),
        "ifRefuses": clean(row[8]),
        "safety": clean(row[9]),
    }

# --------------------------------------------------- protocol, rules, sources
protocols = []
ws = wb["SI NO COME"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if not row or not row[0]:
        continue
    protocols.append({
        "situation": clean(row[0]), "now": clean(row[1]), "never": clean(row[2]),
        "next": clean(row[3]), "when": clean(row[4]), "record": clean(row[5]),
        "level": clean(row[6]), "note": clean(row[7]),
    })

sources = []
ws = wb["FUENTES"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if not row or not row[0]:
        continue
    sources.append({"topic": clean(row[0]), "org": clean(row[1]), "says": clean(row[2]), "url": clean(row[3]), "checked": clean(row[4])})

house_rules, traffic = [], []
ws = wb["INICIO"]
section = None
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    first = clean(row[0]) if row else None
    if first and first.startswith("REGLAS NO NEGOCIABLES"):
        section = "rules"; continue
    if first and first.startswith("SEMÁFORO"):
        section = "traffic"; continue
    if first and first.startswith("IMPORTANTE"):
        disclaimer = first; section = None; continue
    if section == "rules" and first and first.isdigit():
        house_rules.append(clean(row[1]))
    elif section == "traffic" and first in {"VERDE", "AMARILLO", "ROJO"}:
        traffic.append({"light": first, "action": clean(row[1]), "signs": clean(row[2])})

# ------------------------------------------------------------------ emit
os.makedirs(os.path.join(ROOT, "examples/abril"), exist_ok=True)
plan = {
    "generatedFrom": os.path.basename(WORKBOOK),
    "disclaimer": disclaimer,
    "houseRules": house_rules,
    "traffic": traffic,
    "protocols": protocols,
    "sources": sources,
    "days": {},
}

for date in sorted(rows_by_date):
    weekday = weekday_by_date[date]
    steps = sorted(rows_by_date[date], key=lambda r: clock(r[2]))
    planned, guidance = [], {}
    used, nap1_time, nap1_minutes = set(), None, 45
    pending_close = None
    labels_seen = {}
    woke_first = False

    for row in steps:
        time, kind = clock(row[2]), str(row[3]).strip()
        category, priority, minutes, style, before, after = KIND.get(kind, (None,) * 6)

        if kind == "FIN COMIDA":
            pending_close = clean(row[5]) or clean(row[4])   # folded into the meal above
            continue
        # Only the first waking of the day is a step. The other two are naps ending, and a nap that
        # ends is something that happens to you, not something you do -- so it belongs to the nap.
        if kind == "DESPERTAR":
            if woke_first:
                for previous in reversed(planned):
                    if previous["category"] == "sleep" and previous["id"].startswith("siesta"):
                        guidance[previous["name"]]["onWake"] = clean(row[5]) or clean(row[4])
                        break
                continue
            woke_first = True
        if category is None:
            continue

        rule_id = slug(kind)
        n = 2
        while rule_id in used:                              # two blocks of the same kind in a day
            rule_id, n = "%s-%d" % (slug(kind), n), n + 1
        used.add(rule_id)

        rule = {
            "id": rule_id,
            "name": label_for(kind, labels_seen),
            "category": category,
            "priority": priority,
            "durationMinutes": minutes,
        }
        if kind == "SIESTA 1":
            nap1_time, nap1_minutes = time, minutes
        if style == "dependent" and nap1_time:
            # The engine measures a dependency from where the reference *ends*, which is the right
            # question for a baby: what matters is how long she has been awake, not when she fell
            # asleep. So the offset is taken from the end of nap 1, not its start.
            end = int(nap1_time[:2]) * 60 + int(nap1_time[3:]) + nap1_minutes
            gap = (int(time[:2]) * 60 + int(time[3:])) - end
            rule["dependsOn"] = {"rule": "siesta-1", "kind": "dependent",
                                 "offsetMinutes": {"min": gap - 30, "preferred": gap, "max": gap + 15}}
        else:
            rule[("anchor" if style == "anchor" else "window")] = {
                "earliest": shift(time, -before), "preferred": time, "latest": shift(time, after)
            }
        if kind == "CAMA":
            rule["adjustmentPolicies"] = [BEDTIME_POLICY]
        planned.append(rule)

        entry = {
            "time": time,
            "kind": kind,
            "title": clean(row[4]),
            "how": clean(row[5]),
            "food": clean(row[6]),
            "weaning": clean(row[7]),
            "ifRefuses": clean(row[8]),
            "development": clean(row[9]),
            "safety": clean(row[10]),
        }
        menu = menus.get(str(weekday).lower(), {}).get(time)
        if menu:
            entry["menu"] = menu
        guidance[rule["name"]] = {k: v for k, v in entry.items() if v}
        if pending_close and category == "feeding":
            pass
        if pending_close:
            for previous in reversed(planned[:-1]):
                if previous["category"] == "feeding":
                    guidance[previous["name"]]["closing"] = pending_close
                    break
            pending_close = None

    fixture = {
        "date": date,
        "timezone": "America/Vancouver",
        "mode": "normal",
        "currentTime": "07:00",
        "notes": "Generated by scripts/import_routine.py from the family routine workbook. Inputs only.",
        "planned": planned,
        "externalCommitments": [],
        "events": [],
        "manualOverrides": [],
    }
    with open(os.path.join(ROOT, "examples/abril/%s.json" % date), "w", encoding="utf-8") as handle:
        json.dump(fixture, handle, ensure_ascii=False, indent=2)

    plan["days"][date] = {
        "weekday": weekday,
        "stage": stages.get(str(weekday).lower(), {}),
        "guidance": guidance,
    }
    print(date, weekday, len(planned), "reglas")

days_text = {}
for date in sorted(rows_by_date):
    with open(os.path.join(ROOT, "examples/abril/%s.json" % date), encoding="utf-8") as handle:
        days_text[date] = handle.read()

bundle_path = os.path.join(ROOT, "nido-plan.json")
with open(bundle_path, "w", encoding="utf-8") as handle:
    json.dump({"plan": plan, "days": days_text}, handle, ensure_ascii=False)

size = os.path.getsize(bundle_path) / 1024
print("%s: %.0f KB, %d dias, %d protocolos, %d fuentes"
      % (bundle_path, size, len(plan["days"]), len(protocols), len(sources)))
print("Cargalo en la app una vez por telefono. No lo subas a ningun lado.")
